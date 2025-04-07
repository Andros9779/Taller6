import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

# Set a fixed random seed for replicability
np.random.seed(1234)

# ========================================================
# 1. Import the "Data - June 2016" sheet from the Database
# ========================================================
input_filename = "/Users/jdds/Documents/Doing Econ/Taller6/GlobalFinancialDevelopmentDatabaseJune2017.xlsx"
df = pd.read_excel(input_filename, sheet_name="Data - June 2016")

# Print the columns to verify names (adjust if needed)
print("Columnas del DataFrame:", df.columns)

# ========================================================
# 2. Data Cleaning and Winsorization
# ========================================================
#df_clean = df.dropna()
df_clean = df.copy()

def winsorize(series, lower=0.05, upper=0.95):
    lower_val = series.quantile(lower)
    upper_val = series.quantile(upper)
    return series.clip(lower=lower_val, upper=upper_val)

# Update indicator column names as per the dictionary for June 2016.
# For example, we assume the following columns exist:
# "GFDD.DI.12"   -> Private Credit to GDP
# "GFDD.AI.01"   -> Bank Accounts per 1000 adults
# "GFDD.EI.01"   -> Net Interest Margin
# "GFDD.SI.01"   -> Z-Score (Stability)
for col in ["GFDD.DI.12", "GFDD.AI.01", "GFDD.EI.01"]:
    if col in df_clean.columns:
        df_clean[col] = winsorize(df_clean[col])

# ========================================================
# 3. Descriptive Statistics (Overall)
# ========================================================
indicators = ["GFDD.DI.12", "GFDD.AI.01", "GFDD.EI.01", "GFDD.SI.01"]
stats_overall = df_clean[indicators].describe().loc[["mean", "50%", "std", "min", "max"]]
stats_overall.rename(index={"50%": "median"}, inplace=True)

# ========================================================
# 4. Compare Pre‑Crisis and Post‑Crisis Stability (using ZScore)
# ========================================================
# We assume the "Year" column is available.
pre_crisis = df_clean[df_clean["Year"] <= 2007]
post_crisis = df_clean[df_clean["Year"] >= 2009]

stats_pre = pre_crisis["GFDD.SI.01"].describe()
stats_post = post_crisis["GFDD.SI.01"].describe()

# ========================================================
# 5. Visualizations
# ========================================================
# 5a. Scatter Plot: PrivateCredit vs. ZScore (Overall)
plt.figure(figsize=(8,6))
plt.scatter(df_clean["GFDD.DI.12"], df_clean["GFDD.SI.01"], alpha=0.6, color='blue')
plt.xlabel("Private Credit to GDP (%)")
plt.ylabel("Z-Score")
plt.title("Scatter Plot: Private Credit vs. Z-Score (Overall)")
plt.grid(True)
scatter_filename = "scatter_privatecredit_vs_zscore.png"
plt.savefig(scatter_filename)
plt.close()

# 5b. Line Chart: Trend of Average ZScore Over Time
if "Year" in df_clean.columns:
    zscore_trend = df_clean.groupby("Year")["GFDD.SI.01"].mean().reset_index()
    plt.figure(figsize=(8,6))
    plt.plot(zscore_trend["Year"], zscore_trend["GFDD.SI.01"], marker='o', linestyle='-', color='green')
    plt.xlabel("Year")
    plt.ylabel("Average Z-Score")
    plt.title("Trend of Average Z-Score Over Time")
    plt.grid(True)
    line_filename = "line_trend_zscore.png"
    plt.savefig(line_filename)
    plt.close()
else:
    line_filename = None

# 5c. Boxplot: ZScore Pre‑Crisis vs. Post‑Crisis
plt.figure(figsize=(8,6))
plt.boxplot([pre_crisis["GFDD.SI.01"], post_crisis["GFDD.SI.01"]], labels=["Pre-Crisis", "Post-Crisis"])
plt.ylabel("Z-Score")
plt.title("Boxplot of Z-Score: Pre- vs. Post-Crisis")
plt.grid(True)
boxplot_filename = "boxplot_zscore.png"
plt.savefig(boxplot_filename)
plt.close()

# ========================================================
# 6. Pivot Table: Average PrivateCredit by Region
# ========================================================
if "Region" in df_clean.columns:
    pivot = pd.pivot_table(df_clean, values="GFDD.DI.12", index="Region", aggfunc=np.mean)
else:
    pivot = pd.DataFrame()

# ========================================================
# 7. Save All Results into an Excel Workbook
# ========================================================
output_filename = "Financial_Development_Analysis_June2016.xlsx"
with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
    df_clean.to_excel(writer, sheet_name="Clean Data", index=False)
    stats_overall.to_excel(writer, sheet_name="Descriptive Stats Overall")
    pd.DataFrame(stats_pre).to_excel(writer, sheet_name="Pre-Crisis ZScore Stats")
    pd.DataFrame(stats_post).to_excel(writer, sheet_name="Post-Crisis ZScore Stats")
    if not pivot.empty:
        pivot.to_excel(writer, sheet_name="Pivot: PrivateCredit by Region")
    
    workbook = writer.book
    from openpyxl.drawing.image import Image as XLImage
    worksheet1 = workbook.create_sheet("Scatter Plot")
    worksheet1.add_image(XLImage(scatter_filename), "B2")
    if line_filename:
        worksheet2 = workbook.create_sheet("Line Chart")
        worksheet2.add_image(XLImage(line_filename), "B2")
    worksheet3 = workbook.create_sheet("Boxplot ZScore")
    worksheet3.add_image(XLImage(boxplot_filename), "B2")

print(f"Excel file '{output_filename}' has been created with all analyses.")